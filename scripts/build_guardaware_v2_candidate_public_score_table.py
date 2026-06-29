from pathlib import Path
import csv
import json
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except Exception as e:
    raise SystemExit(
        "openpyxl is required to create .xlsx. "
        "Install it with: python -m pip install openpyxl"
    ) from e


OUT_DIR = Path("analysis/public_benchmark_eval/guardaware_v2_candidate_public_benchmark_run1")
SUMMARY_CSV = OUT_DIR / "candidate_public_benchmark_score_summary.csv"
XLSX = OUT_DIR / "guardaware_v2_candidate_public_score_table.xlsx"

CANDIDATE_CKPT = "checkpoints/probes/15x15_b4c64_guardaware_v2_modefix_candidate_run1.pt"
C_WEIGHTS = "c_inference/weights/15x15_b4c64_guardaware_v2_modefix_candidate_run1_weights.bin"

# From the user's pre-change public benchmark score table.
OLD_CURRENT_BEST = {
    "random": 24.0,
    "tactical_lite": 23.0,
    "tactical_mid": 7.0,
    "tactical_plus": 3.0,
    "rapfi_fast_depth1": 0.0,
}

TAKEAWAYS = {
    "random": "Solved; no change vs previous model.",
    "tactical_lite": "Near solved; no change vs previous model.",
    "tactical_mid": "Major tactical weakness remains; no public benchmark gain.",
    "tactical_plus": "Strong tactical baseline still exposes weakness; no public benchmark gain.",
    "rapfi_fast_depth1": "Still cannot beat shallow Rapfi; no public benchmark gain.",
}

DISPLAY_NAMES = {
    "random": "random",
    "tactical_lite": "tactical_lite",
    "tactical_mid": "tactical_mid",
    "tactical_plus": "tactical_plus",
    "rapfi_fast_depth1": "rapfi_fast_depth1",
}

rows = []
with SUMMARY_CSV.open(newline="") as f:
    reader = csv.DictReader(f)
    for r in reader:
        baseline = r["baseline_opponent"]
        candidate_score = float(r["candidate_score"])
        old_score = OLD_CURRENT_BEST[baseline]
        rows.append({
            "baseline_opponent": DISPLAY_NAMES.get(baseline, baseline),
            "rapfi_full_score": float(r["rapfi_full_score_reference"]),
            "neural_current_best_before_score": old_score,
            "guardaware_v2_candidate_score": candidate_score,
            "delta_vs_current_best": candidate_score - old_score,
            "candidate_config": r["candidate_config"],
            "candidate_wld": f'{r["candidate_wins"]}-{r["candidate_losses"]}-{r["candidate_draws"]}',
            "candidate_score_rate": float(r["candidate_score_rate"]),
            "gap_vs_rapfi_full": float(r["gap_vs_rapfi_full"]),
            "parse_status": r["parse_status"],
            "log_file": r["log_file"],
            "takeaway": TAKEAWAYS.get(baseline, ""),
        })

wb = Workbook()
ws = wb.active
ws.title = "Score Ladder"

meta = wb.create_sheet("Run Metadata")
evid = wb.create_sheet("Evidence Logs")

# Theme
navy = "1F4E78"
blue = "D9EAF7"
green = "E2F0D9"
yellow = "FFF2CC"
red = "FCE4D6"
gray = "F2F2F2"
white = "FFFFFF"
thin_gray = Side(style="thin", color="D9D9D9")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

# Score Ladder sheet
ws["A1"] = "Public benchmark score ladder, 15x15 freestyle, 24 games each"
ws["A1"].font = Font(bold=True, size=15, color=white)
ws["A1"].fill = PatternFill("solid", fgColor=navy)
ws.merge_cells("A1:L1")

ws["A2"] = "Comparison: before capacity/data increase current-best vs guard-aware v2 saved candidate"
ws["A2"].font = Font(italic=True, color="666666")
ws.merge_cells("A2:L2")

headers = [
    "Baseline opponent",
    "Rapfi full score",
    "Neural current-best before",
    "Guard-aware v2 candidate",
    "Delta vs current-best",
    "Candidate config",
    "Candidate W-L-D",
    "Candidate score rate",
    "Gap vs Rapfi",
    "Parse status",
    "Takeaway",
    "Log file",
]
ws.append([])
ws.append(headers)

header_row = 4
for cell in ws[header_row]:
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for r in rows:
    ws.append([
        r["baseline_opponent"],
        r["rapfi_full_score"],
        r["neural_current_best_before_score"],
        r["guardaware_v2_candidate_score"],
        r["delta_vs_current_best"],
        r["candidate_config"],
        r["candidate_wld"],
        r["candidate_score_rate"],
        r["gap_vs_rapfi_full"],
        r["parse_status"],
        r["takeaway"],
        r["log_file"],
    ])

for row in range(5, 5 + len(rows)):
    for col in range(1, 13):
        c = ws.cell(row=row, column=col)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
    delta = ws.cell(row=row, column=5).value
    if delta > 0:
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=green)
    elif delta < 0:
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=red)
    else:
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=yellow)

    if ws.cell(row=row, column=10).value == "OK":
        ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=green)
    else:
        ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=red)

for row in range(5, 5 + len(rows)):
    ws.cell(row=row, column=8).number_format = "0.0%"
    ws.cell(row=row, column=2).number_format = "0.0"
    ws.cell(row=row, column=3).number_format = "0.0"
    ws.cell(row=row, column=4).number_format = "0.0"
    ws.cell(row=row, column=5).number_format = "+0.0;-0.0;0.0"
    ws.cell(row=row, column=9).number_format = "+0.0;-0.0;0.0"

summary_start = 11
ws[f"A{summary_start}"] = "Summary verdict"
ws[f"A{summary_start}"].font = Font(bold=True, color=white)
ws[f"A{summary_start}"].fill = PatternFill("solid", fgColor=navy)
ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=12)

summary_lines = [
    ("Overall result", "No public benchmark improvement over the previous current-best score table."),
    ("Best observed candidate score", f'{max(r["guardaware_v2_candidate_score"] for r in rows):.1f}/24'),
    ("Worst remaining gap", f'{min(r["gap_vs_rapfi_full"] for r in rows):.1f} vs Rapfi full score'),
    ("Promotion status", "Not promoted. Candidate only; current-best was not overwritten."),
    ("Use in report", "This Excel can be used as the post-training/data-increase public benchmark score table."),
]
for i, (k, v) in enumerate(summary_lines, start=summary_start + 1):
    ws.cell(row=i, column=1).value = k
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=2).value = v
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
    for col in range(1, 13):
        ws.cell(row=i, column=col).border = border
        ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")

# Metadata sheet
meta["A1"] = "Run metadata"
meta["A1"].font = Font(bold=True, size=15, color=white)
meta["A1"].fill = PatternFill("solid", fgColor=navy)
meta.merge_cells("A1:D1")

metadata_rows = [
    ("Generated at", datetime.now().isoformat(timespec="seconds")),
    ("Candidate checkpoint", CANDIDATE_CKPT),
    ("Candidate C weights", C_WEIGHTS),
    ("Candidate route", "b4c64 guard-aware v2 modefix saved candidate run1"),
    ("Board size", 15),
    ("Win length", 5),
    ("Channels", 64),
    ("Blocks", 4),
    ("Training route status", "PASS-gated saved candidate ready for public benchmark"),
    ("Public benchmark games", "24 games per baseline opponent"),
    ("Current-best overwrite", "No"),
    ("Promotion", "No"),
    ("C export", "Yes, for benchmark scoring only"),
    ("Excel conclusion", "Candidate matches old current-best score ladder; no benchmark improvement."),
]
meta.append([])
meta.append(["Field", "Value"])
for c in meta[3]:
    c.font = Font(bold=True, color=white)
    c.fill = PatternFill("solid", fgColor=navy)
    c.border = border
for k, v in metadata_rows:
    meta.append([k, v])
for row in meta.iter_rows(min_row=4, max_row=3 + len(metadata_rows), min_col=1, max_col=2):
    for c in row:
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
    row[0].font = Font(bold=True)

# Evidence sheet
evid["A1"] = "Evidence logs"
evid["A1"].font = Font(bold=True, size=15, color=white)
evid["A1"].fill = PatternFill("solid", fgColor=navy)
evid.merge_cells("A1:D1")
evid.append([])
evid.append(["Baseline", "Config", "Log file", "Parse status"])
for c in evid[3]:
    c.font = Font(bold=True, color=white)
    c.fill = PatternFill("solid", fgColor=navy)
    c.border = border
for r in rows:
    evid.append([r["baseline_opponent"], r["candidate_config"], r["log_file"], r["parse_status"]])
for row in evid.iter_rows(min_row=4, max_row=3 + len(rows), min_col=1, max_col=4):
    for c in row:
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)

# Formatting
widths = {
    "A": 22, "B": 16, "C": 22, "D": 24, "E": 20, "F": 16,
    "G": 16, "H": 18, "I": 16, "J": 14, "K": 42, "L": 70,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for row in range(1, 5 + len(rows)):
    ws.row_dimensions[row].height = 24
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:L{4 + len(rows)}"

for sheet in [meta, evid]:
    for col in range(1, sheet.max_column + 1):
        letter = get_column_letter(col)
        if letter == "A":
            sheet.column_dimensions[letter].width = 26
        elif letter == "B":
            sheet.column_dimensions[letter].width = 80
        elif letter == "C":
            sheet.column_dimensions[letter].width = 80
        else:
            sheet.column_dimensions[letter].width = 18
    sheet.freeze_panes = "A4"

# Add a tiny JSON summary next to the xlsx for git-friendly review.
json_summary = {
    "xlsx": str(XLSX),
    "generated_at": datetime.now().isoformat(timespec="seconds"),
    "candidate_checkpoint": CANDIDATE_CKPT,
    "candidate_c_weights": C_WEIGHTS,
    "rows": rows,
    "conclusion": "No public benchmark improvement vs previous current-best table.",
}
(OUT_DIR / "guardaware_v2_candidate_public_score_table_summary.json").write_text(
    json.dumps(json_summary, indent=2, sort_keys=True) + "\n"
)

wb.save(XLSX)
print("wrote", XLSX)
print("wrote", OUT_DIR / "guardaware_v2_candidate_public_score_table_summary.json")
